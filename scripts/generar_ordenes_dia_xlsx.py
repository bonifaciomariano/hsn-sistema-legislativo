"""Regenera ordenes_dia_2026.xlsx desde data/od.json + data/proyectos.json.

ordenes_dia_2026.xlsx es la fuente primaria que lee parse_ayuda_memoria.py
(ver ese script). Hasta ahora era un export manual del buscador de Órdenes
del Día del Senado, y había quedado desactualizado (tope OD 272) mientras
data/od.json -- que sí se scrapea automáticamente -- ya tenía hasta OD 361.
Este script cierra esa brecha generando el xlsx con el mismo esquema de
columnas que ya espera leer_ordenes_dia(), a partir de datos que el repo ya
tiene scrapeados (no hace falta bajar ningún PDF nuevo).

Incluye todo OD 2026 (tipo NORMAL o ANEXO) cuyo expediente NO figure ya
aprobado/sancionado en data/sanciones.json (Boletín de Novedades) -- Ayuda
Memoria es para lo que todavía no se trató, no un archivo histórico
completo.

Ejecutar con: py scripts/generar_ordenes_dia_xlsx.py
Después: py scripts/parse_ayuda_memoria.py (cruza contra
AYUDA_MEMORIA_2026.xlsx, que Mariano mantiene a mano con firmantes/comisión
cabecera -- este script no lo toca).
"""
import json
import os

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE, "data")
OUT_PATH = os.path.join(BASE, "ordenes_dia_2026.xlsx")


def _cargar(nombre):
    with open(os.path.join(DATA_DIR, nombre), encoding="utf-8") as f:
        return json.load(f)


def _codigo_exp(e):
    return f"{e['origen']}-{e['nro']}/{str(e['anio'])[-2:]}-{e['tipo']}"


def _autor_extracto(p):
    """Arma el texto crudo tal como lo espera _parsear_extracto() de
    parse_ayuda_memoria.py: 'APELLIDO [Y OTROS]: descripcion' o, si es
    mensaje del PE (sin autores individuales), la descripción sola
    (empieza con 'MENSAJE')."""
    extracto = p.get("extracto") or ""
    autores = p.get("autores") or []
    if not autores:
        return extracto
    apellido = autores[0].split(",")[0].strip()
    prefijo = apellido if len(autores) == 1 else f"{apellido} Y OTROS"
    return f"{prefijo}: {extracto}"


def main():
    od = _cargar("od.json")
    proyectos = _cargar("proyectos.json")
    sanciones = _cargar("sanciones.json")

    pidx = {(p["origen"], p["nro"], p["anio"]): p for p in proyectos}

    tratados = set()
    for s in sanciones:
        if s["seccion"] in ("ley", "decreto_res_com_dec", "acuerdo") and \
                (s.get("resultado") or "").upper() in ("APROBADO", "APROBADA"):
            tratados.add(s["expediente"])

    od2026 = [o for o in od if o.get("anio_od") == 2026]

    filas = []
    for o in sorted(od2026, key=lambda x: (x["nro_od"], x["tipo_od"])):
        exps = o.get("expedientes") or []
        if not exps:
            continue
        codigos_simples = [f"{e['origen']}-{e['nro']}/{str(e['anio'])[-2:]}" for e in exps]
        if any(c in tratados for c in codigos_simples):
            continue  # ya tratado según Boletín de Novedades -- no va en Ayuda Memoria

        primer_exp = exps[0]
        p = pidx.get((primer_exp["origen"], primer_exp["nro"], primer_exp["anio"]))
        if not p:
            continue

        giros = " - ".join(p.get("comisiones") or []) + (" -" if p.get("comisiones") else "")
        filas.append({
            "numero": o["nro_od"],
            "tipo": "ANEXO" if o["tipo_od"] == "A" else "NORMAL",
            "codigos": [_codigo_exp(e) for e in exps],
            "exp_url": p.get("url"),
            "extracto": _autor_extracto(p),
            "giros": giros,
            "od_url": o.get("url_pdf"),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Órdenes del Día"
    ws.append(["Periodo", "Número", "Tipo", "Sobre los expedientes", "Extracto",
               "Giros", "Fecha Dictamen", "Estado", "Adjunto"])

    for f in filas:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=2026)
        ws.cell(row=r, column=2, value=f["numero"])
        ws.cell(row=r, column=3, value=f["tipo"])
        c4 = ws.cell(row=r, column=4, value="\n".join(f["codigos"]))
        if f["exp_url"]:
            c4.hyperlink = f["exp_url"]
        ws.cell(row=r, column=5, value=f["extracto"])
        ws.cell(row=r, column=6, value=f["giros"])
        ws.cell(row=r, column=7, value=None)
        ws.cell(row=r, column=8, value="PE")
        c9 = ws.cell(row=r, column=9, value="Descargar" if f["od_url"] else None)
        if f["od_url"]:
            c9.hyperlink = f["od_url"]

    wb.save(OUT_PATH)
    print(f"Total filas: {len(filas)}")
    print(f"Guardado en {OUT_PATH}")


if __name__ == "__main__":
    main()
