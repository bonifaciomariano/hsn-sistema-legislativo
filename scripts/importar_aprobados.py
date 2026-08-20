#!/usr/bin/env python3
"""Refresca sancionado/ley_numero/fecha_ley/archivado/... en data/proyectos.json
a partir de aprobados_2025.xlsx + aprobados_2026.xlsx — export de expedientes
aprobados por el HSN (mismo esquema de 39 columnas que 2025.xlsx/2026.xlsx de
la migración inicial, sólo que pre-filtrado a aprobados).

A diferencia de esos xlsx, acá el encabezado no está en la fila 1: filas 1-3
son título/fecha del reporte, fila 4 vacía, fila 5 el encabezado real
("ORIGEN, NRO., AÑO, ..."), fila 6 vacía, y recién de la fila 7 en adelante
los datos — por eso no se reusa leer_filas() de importar_proyectos_xlsx.py
tal cual, sólo parsear_fila()/cargar_padron().

Sólo actualiza los campos de trazabilidad de sanción/archivo de expedientes
YA existentes en proyectos.json (por clave nro+año+tipo) — no toca autores,
bloques, comisiones, etc., que los sigue manteniendo scraper_proyectos.py. Si
un expediente no está todavía en proyectos.json, se inserta completo.

Ejecutar con: python scripts/importar_aprobados.py
"""
import json
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from importar_proyectos_xlsx import parsear_fila  # noqa: E402
from scraper_proyectos import cargar_padron  # noqa: E402

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(REPO_ROOT, "data")
PROYECTOS_JSON = os.path.join(DATA_DIR, "proyectos.json")

APROBADOS_PATHS = [
    os.path.join(REPO_ROOT, "aprobados_2025.xlsx"),
    os.path.join(REPO_ROOT, "aprobados_2026.xlsx"),
]

CAMPOS_TRAZABILIDAD = [
    "sancionado", "ley_numero", "fecha_ley",
    "archivado", "fecha_archivo", "caduca", "fecha_caduca",
]


def leer_filas_aprobados(path):
    """El encabezado real ('ORIGEN,...') está en la primera fila no vacía de
    las primeras 10 — se detecta en vez de asumir una posición fija, porque
    estos reportes traen 3-4 filas de título antes."""
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and row[0] == "ORIGEN":
            header_row = i
            break
    if header_row is None:
        raise RuntimeError(f"No se encontró la fila de encabezados ('ORIGEN') en {path}")
    filas = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[1]:  # columna NRO. (fórmula HYPERLINK) vacía = fin de datos
            continue
        filas.append(row)
    return filas


def main():
    padron, indice = cargar_padron()
    proyectos = json.load(open(PROYECTOS_JSON, encoding="utf-8"))
    indice_proyectos = {(p["nro"], p["anio"], p["tipo"]): p for p in proyectos}

    actualizados, insertados, sin_cambios = 0, 0, 0
    for path in APROBADOS_PATHS:
        if not os.path.exists(path):
            print(f"AVISO: no se encontró {path}, se omite")
            continue
        filas = leer_filas_aprobados(path)
        print(f"  {os.path.basename(path)}: {len(filas)} filas")
        for row in filas:
            tipo = (row[3] or "").strip()
            if not tipo:
                continue
            nuevo = parsear_fila(row, padron, indice)
            clave = (nuevo["nro"], nuevo["anio"], nuevo["tipo"])
            existente = indice_proyectos.get(clave)
            if existente is None:
                proyectos.append(nuevo)
                indice_proyectos[clave] = nuevo
                insertados += 1
                continue
            cambio = False
            for campo in CAMPOS_TRAZABILIDAD:
                if existente.get(campo) != nuevo.get(campo):
                    existente[campo] = nuevo.get(campo)
                    cambio = True
            if cambio:
                actualizados += 1
            else:
                sin_cambios += 1

    proyectos.sort(key=lambda p: (int(p["anio"]), int(p["nro"])), reverse=True)
    json.dump(proyectos, open(PROYECTOS_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    print(f"  → {actualizados} proyectos actualizados, {insertados} insertados, "
          f"{sin_cambios} ya estaban al día")
    sancionados = sum(1 for p in proyectos if p.get("sancionado"))
    print(f"  → TOTAL sancionados (convertidos en ley) en proyectos.json: {sancionados}")


if __name__ == "__main__":
    main()
