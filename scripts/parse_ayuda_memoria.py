"""Migra los xlsx manuales de Ayuda Memoria a data/ayuda_memoria.json.

Fuente primaria: ordenes_dia_2025.xlsx + ordenes_dia_2026.xlsx (export
"Órdenes del Día" del Senado, uno por período; son los más actualizados y ya
traen expedientes/giros/fechas estructurados e hipervínculos reales, pero sin
firmantes). El de 2025 se filtra a sólo OD de Proyecto de Ley (PL) — a
pedido de Mariano, el resto de las categorías de ese período ya no interesa
para Ayuda Memoria.

Fuente de enriquecimiento: AYUDA_MEMORIA_2026.xlsx, hojas "OD LEY",
"ANEXO I" y "OD ACUERDOS" (las únicas 3 que corresponden a esta sección; se
ignoran Índice, "AC para dar cuenta", Anexos II/IV/V y "PL del PEN"). De ahí
se toman los FIRMANTES del dictamen de mayoría y, para OD LEY/ANEXO I, la
columna COMISION CABECERA.

Ejecutar con: py scripts/parse_ayuda_memoria.py

Nota para cuando vuelva a andar el scraper del sitio del Senado (bloqueado
por un anti-bot desde el 5/8/2026): este script deja de tener sentido y el
cruce firmantes/comisión cabecera debería salir directo del scrape del
expediente, igual que se documentó en importar_acuerdos.py para "dado
cuenta".
"""
import json
import os
import re
import unicodedata
from datetime import datetime

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# (ruta, categoría a la que se restringe ese período — None = todas)
ORDENES_DIA_FUENTES = [
    (os.path.join(BASE, "ordenes_dia_2025.xlsx"), "PL"),
    (os.path.join(BASE, "ordenes_dia_2026.xlsx"), None),
]
AYUDA_MEMORIA_PATH = os.path.join(BASE, "AYUDA_MEMORIA_2026.xlsx")
JSON_PATH = os.path.join(BASE, "data", "ayuda_memoria.json")

ORIGEN_MAP = {"PE": "Poder Ejecutivo", "S": "Senado", "CD": "Cámara de Diputados"}
CATEGORIA_MAP = {
    "PL": "Proyecto de Ley",
    "PD": "Proyecto de Declaración",
    "PC": "Proyecto de Comunicación",
    "PR": "Proyecto de Resolución",
    "AC": "Acuerdo",
}
PARTICULAS = {"DE", "DEL", "LA", "LOS", "LAS", "Y", "E"}
# Conectores gramaticales que se dejan en minúscula al title-casear la
# descripción (salvo que sean la primera palabra de la oración): evita que
# "solicita", "que", "para" etc. queden en mayúscula, sin necesidad de
# reconocer qué palabras son nombres propios (el texto fuente viene todo en
# mayúsculas, así que no hay forma de distinguirlos por caja).
CONECTORES = PARTICULAS | {
    "A", "EL", "EN", "QUE", "POR", "PARA", "CON", "SU", "SUS", "AL", "UN",
    "UNA", "UNOS", "UNAS", "SOBRE", "ENTRE", "SE", "NO", "LO", "COMO", "SIN",
    "TRAS",
}

RE_EXP_CODIGO = re.compile(r"^([A-Z.]+)-\d+/\d+(?:-([A-Z]{2}))?$")


# ── Helpers de texto ────────────────────────────────────────────────────────

def _sin_tildes(s):
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _titlecase_particulas(raw):
    """'CAPITANICH Y OTROS' -> 'Capitanich y otros'; conserva partículas
    (de/del/la/y) en minúscula salvo que sean la primera palabra."""
    palabras = raw.split()
    out = []
    for i, w in enumerate(palabras):
        if not w:
            continue
        if w.upper() in PARTICULAS and i != 0:
            out.append(w.lower())
        else:
            out.append(w[0].upper() + w[1:].lower())
    return " ".join(out)


RE_ORDINAL = re.compile(r"^\d+[a-zñ]{1,3}$")


def _capitalizar(palabra):
    """Pone en mayúscula la primera letra alfabética de la palabra, dejando
    intacta cualquier puntuación/número pegado (paréntesis, comas, '24.898').
    No toca ordinales tipo '8va'/'1ro' (mayuscular la letra pegada al número
    quedaría '8Va')."""
    if RE_ORDINAL.match(palabra):
        return palabra
    m = re.search(r"[a-zA-Zá-úñÁ-ÚÑ]", palabra)
    if not m:
        return palabra
    i = m.start()
    return palabra[:i] + palabra[i].upper() + palabra[i + 1:]


def sentence_case(texto):
    """Texto en MAYÚSCULAS -> Title Case (cada palabra con inicial
    mayúscula), dejando en minúscula los conectores gramaticales salvo que
    empiecen oración. No intenta reconocer nombres propios/siglas
    específicos (el texto fuente viene todo en mayúsculas, no hay señal de
    caja para distinguirlos) — el trade-off es capitalizar también verbos
    comunes, pero es preferible a que los nombres propios y siglas queden
    en minúscula."""
    t = (texto or "").strip()
    if not t:
        return t
    palabras = t.lower().split(" ")
    out = []
    inicio_oracion = True
    for w in palabras:
        limpio = re.sub(r"[^\wÁÉÍÓÚÑáéíóúñ]", "", w, flags=re.UNICODE)
        clave = _sin_tildes(limpio).upper()
        if inicio_oracion or (clave and clave not in CONECTORES):
            out.append(_capitalizar(w))
        else:
            out.append(w)
        inicio_oracion = w.rstrip().endswith((".", "!", "?"))
    return " ".join(out)


def _limpiar_giros(giros_raw):
    """'DE X - DE Y -' -> ['De X', 'De Y'] sin duplicados, orden estable."""
    if not giros_raw:
        return []
    partes = [p.strip(" -") for p in giros_raw.split(" - ")]
    vistos = []
    for p in partes:
        if not p:
            continue
        bonito = _titlecase_particulas(p)
        if bonito not in vistos:
            vistos.append(bonito)
    return vistos


def _origen_de_codigo(codigo):
    m = re.match(r"^([A-Z.]+)-", codigo.strip().upper())
    prefijo = m.group(1).replace(".", "") if m else ""
    return ORIGEN_MAP.get(prefijo, prefijo.title() if prefijo else "")


def _categoria_de_codigo(codigo):
    m = RE_EXP_CODIGO.match(codigo.strip().upper())
    tipo = m.group(2) if m and m.group(2) else ""
    return CATEGORIA_MAP.get(tipo, "")


# ── ordenes_dia_AAAA.xlsx (fuente primaria, una por período) ────────────────

def _tipo_de_codigo(codigo):
    """'PE-231/26-AC' -> 'AC'. Tipo crudo del expediente (no el label de
    CATEGORIA_MAP), para poder filtrar por período sin depender de qué
    categorías estén mapeadas a un label lindo."""
    m = RE_EXP_CODIGO.match((codigo or "").strip().upper())
    return m.group(2) if m and m.group(2) else ""


def leer_ordenes_dia(path, filtro_tipo=None):
    """filtro_tipo: si se da (ej. 'PL'), sólo se incluyen filas cuyo primer
    expediente sea de ese tipo — se usa para restringir ordenes_dia_2025.xlsx
    a sólo Proyectos de Ley."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = []
    for r in range(2, ws.max_row + 1):
        periodo = ws.cell(row=r, column=1).value
        numero = ws.cell(row=r, column=2).value
        if periodo is None or numero is None:
            continue
        tipo_cell = ws.cell(row=r, column=3).value or "NORMAL"
        tipo_od = "ANEXO" if str(tipo_cell).strip().upper().startswith("ANEXO") else "NORMAL"
        exp_cell = ws.cell(row=r, column=4)
        exp_raw = (exp_cell.value or "").strip()
        codigos = [c.strip() for c in re.split(r"[\n]| - ", exp_raw) if c.strip()]
        exp_link = exp_cell.hyperlink.target if exp_cell.hyperlink else None
        expedientes = [
            {"codigo": c, "url": exp_link if i == 0 else None}
            for i, c in enumerate(codigos)
        ]
        primer_codigo = codigos[0] if codigos else ""
        if filtro_tipo and _tipo_de_codigo(primer_codigo) != filtro_tipo:
            continue

        extracto_raw = (ws.cell(row=r, column=5).value or "").strip()
        giros_raw = ws.cell(row=r, column=6).value
        fecha = ws.cell(row=r, column=7).value
        adjunto_cell = ws.cell(row=r, column=9)
        od_link = adjunto_cell.hyperlink.target if adjunto_cell.hyperlink else None

        autor, descripcion = _parsear_extracto(extracto_raw)

        filas.append({
            "numero": str(int(numero)) if isinstance(numero, (int, float)) else str(numero).strip(),
            "periodo": int(periodo),
            "tipoOD": tipo_od,
            "origen": _origen_de_codigo(primer_codigo) if primer_codigo else "",
            "categoria": _categoria_de_codigo(primer_codigo) if primer_codigo else "",
            "autor": autor,
            "descripcion": descripcion,
            "expedientes": expedientes,
            "comisiones": _limpiar_giros(giros_raw),
            "comisionCabecera": None,
            "fechaDictamen": fecha if isinstance(fecha, str) else (fecha.strftime("%d/%m/%Y") if fecha else None),
            "odLink": od_link,
            "firmantesMayoria": None,
        })
    return filas


def _parsear_extracto(extracto_raw):
    """'CAPITANICH Y OTROS: PROYECTO...' -> ('Capitanich y otros', 'Proyecto...').
    'MENSAJE N° ...' (sin autor individual, mensaje del PEN) -> (None, texto)."""
    if not extracto_raw:
        return None, ""
    if extracto_raw.upper().startswith("MENSAJE"):
        return None, sentence_case(extracto_raw)
    idx = extracto_raw.find(":")
    if 0 < idx <= 60:
        autor_raw = extracto_raw[:idx].strip()
        resto = extracto_raw[idx + 1:].strip()
        return _titlecase_particulas(autor_raw), sentence_case(resto)
    return None, sentence_case(extracto_raw)


# ── AYUDA_MEMORIA_2026.xlsx (enriquecimiento: firmantes + comisión cabecera) ─

def _parsear_od_numero(valor, formato_celda):
    """'275/2026 (N)' -> (275, 2026, 'NORMAL'). Corrige el caso en que Excel
    interpretó '5/2026' como fecha (datetime 2026-05-01, formato m/yyyy)."""
    if isinstance(valor, datetime):
        return valor.month, valor.year, "NORMAL"
    s = str(valor or "").strip()
    tipo = "ANEXO" if "(A)" in s.upper() else "NORMAL"
    s = re.sub(r"\s*\([AN]\)\s*", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+y\s+anexo.*$", "", s, flags=re.IGNORECASE).strip()
    m = re.match(r"^(\d+)\s*/\s*(\d+)$", s)
    if not m:
        return None, None, tipo
    nro = int(m.group(1))
    anio = int(m.group(2))
    if anio < 100:
        anio += 2000
    return nro, anio, tipo


def _parsear_firmantes(texto):
    if not texto:
        return []
    partes = re.split(r"\s+–\s+|\s+-\s+", texto.strip())
    out = []
    for p in partes:
        p = p.strip().rstrip(".").strip()
        if p:
            out.append(p)
    return out


def leer_enriquecimiento():
    """-> dict (numero, periodo, tipoOD) -> {firmantes, comisionCabecera}."""
    wb = openpyxl.load_workbook(AYUDA_MEMORIA_PATH, data_only=True)
    resultado = {}
    hojas = [("OD LEY", True), ("ANEXO I", True), ("OD ACUERDOS", False)]
    for nombre_hoja, tiene_comision in hojas:
        if nombre_hoja not in wb.sheetnames:
            continue
        ws = wb[nombre_hoja]
        for r in range(4, ws.max_row + 1):
            od_valor = ws.cell(row=r, column=1).value
            if od_valor is None or str(od_valor).strip() == "":
                continue
            firmantes_txt = ws.cell(row=r, column=4).value or ""
            comision = ws.cell(row=r, column=5).value if tiene_comision else None

            # Caso especial: '113/2026 y ANEXO' trae firmantes del dictamen
            # principal + 'ANEXO: ...' con firmantes adicionales -> se
            # reparten entre la fila NORMAL y la fila ANEXO de ese mismo OD.
            if isinstance(od_valor, str) and "y anexo" in od_valor.lower():
                nro, anio, _ = _parsear_od_numero(od_valor, None)
                principal, _, anexo = firmantes_txt.partition("ANEXO:")
                if nro is not None:
                    resultado[(nro, anio, "NORMAL")] = {
                        "firmantes": _parsear_firmantes(principal),
                        "comisionCabecera": comision,
                    }
                    if anexo.strip():
                        resultado[(nro, anio, "ANEXO")] = {
                            "firmantes": _parsear_firmantes(anexo),
                            "comisionCabecera": comision,
                        }
                continue

            nro, anio, tipo = _parsear_od_numero(od_valor, None)
            if nro is None:
                continue
            resultado[(nro, anio, tipo)] = {
                "firmantes": _parsear_firmantes(firmantes_txt),
                "comisionCabecera": comision,
            }
    return resultado


def main():
    filas = []
    for path, filtro_tipo in ORDENES_DIA_FUENTES:
        if not os.path.exists(path):
            print(f"AVISO: no se encontró {path}, se omite esa fuente")
            continue
        nuevas = leer_ordenes_dia(path, filtro_tipo)
        print(f"  {os.path.basename(path)}: {len(nuevas)} filas"
              + (f" (filtradas a tipo {filtro_tipo})" if filtro_tipo else ""))
        filas.extend(nuevas)
    enriquecimiento = leer_enriquecimiento()

    matcheados = 0
    for f in filas:
        try:
            nro_int = int(f["numero"])
        except ValueError:
            continue
        clave = (nro_int, f["periodo"], f["tipoOD"])
        extra = enriquecimiento.get(clave)
        if not extra:
            continue
        matcheados += 1
        if extra["firmantes"]:
            f["firmantesMayoria"] = extra["firmantes"]
        if extra["comisionCabecera"]:
            f["comisionCabecera"] = extra["comisionCabecera"]
        elif f["categoria"] == "Acuerdo":
            f["comisionCabecera"] = "Acuerdos"

    for f in filas:
        if f["categoria"] == "Acuerdo" and not f["comisionCabecera"]:
            f["comisionCabecera"] = "Acuerdos"

    filas.sort(key=lambda f: (f["periodo"], int(f["numero"]) if f["numero"].isdigit() else 0), reverse=True)

    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as fh:
        json.dump(filas, fh, ensure_ascii=False, indent=2)

    con_firmantes = sum(1 for f in filas if f["firmantesMayoria"])
    print(f"Total: {len(filas)} órdenes del día")
    print(f"Cruzados con Ayuda Memoria (firmantes/comisión): {matcheados}")
    print(f"Con firmantes: {con_firmantes}")
    print(f"Sin firmantes: {len(filas) - con_firmantes}")


if __name__ == "__main__":
    main()
