#!/usr/bin/env python3
"""
construir_senadores.py — Construye data/senadores.json (padrón con bloques/provincias)
========================================================================================
Lee los dos Excel del repo:
    Senadores_2026.xlsx                  → senadores en ejercicio (BLOQUE, APELLIDO, NOMBRE, PROVINCIA, ...)
    Senadores_mandato_cumplido_2025.xlsx → mandatos cumplidos (Nombre, Distrito, BLOQUE)  [fallback]

Genera un padrón:
    {
      "APELLIDO, NOMBRE": {"bloque": "...", "provincia": "...", "vigente": true},
      ...
    }

Se importa desde scraper_proyectos.py; también puede ejecutarse directo para regenerar
data/senadores.json.
"""

import json
import os
import unicodedata

import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(REPO_ROOT, "data")
XLSX_VIGENTES = os.path.join(REPO_ROOT, "Senadores_2026.xlsx")
XLSX_CUMPLIDOS = os.path.join(REPO_ROOT, "Senadores_mandato_cumplido_2025.xlsx")
SALIDA = os.path.join(DATA_DIR, "senadores.json")


def normalizar_nombre(texto):
    """'  Blanco, Pablo Daniel ' → 'BLANCO, PABLO DANIEL'"""
    if not texto:
        return ""
    texto = " ".join(str(texto).split()).strip()
    if "," in texto:
        apellido, nombre = texto.split(",", 1)
        texto = f"{apellido.strip()}, {nombre.strip()}"
    return texto.upper()


def sin_tildes(texto):
    return "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )


def construir_padron():
    padron = {}

    # 1) Senadores en ejercicio (xlsx 2026)
    if os.path.exists(XLSX_VIGENTES):
        wb = openpyxl.load_workbook(XLSX_VIGENTES, read_only=True)
        ws = wb.active
        for fila in list(ws.iter_rows(values_only=True))[1:]:
            bloque, apellido, nombre, provincia = (fila[0], fila[1], fila[2], fila[3])
            if not apellido:
                continue
            nombre_norm = normalizar_nombre(f"{apellido}, {nombre}" if nombre else apellido)
            padron[nombre_norm] = {
                "bloque": (bloque or "Sin datos").strip(),
                "provincia": (provincia or "").strip(),
                "vigente": True,
            }
        wb.close()

    # 2) Mandatos cumplidos (xlsx 2025) — fallback, no pisa a los vigentes
    if os.path.exists(XLSX_CUMPLIDOS):
        wb = openpyxl.load_workbook(XLSX_CUMPLIDOS, read_only=True)
        ws = wb.active
        for fila in list(ws.iter_rows(values_only=True))[1:]:
            nombre_raw, distrito, bloque = (fila[0], fila[1], fila[2])
            if not nombre_raw:
                continue
            nombre_norm = normalizar_nombre(nombre_raw)
            if nombre_norm in padron:
                continue
            padron[nombre_norm] = {
                "bloque": (bloque or "Sin datos").strip(),
                "provincia": (distrito or "").strip().upper(),
                "vigente": False,
            }
        wb.close()

    return padron


def construir_indice_apellido(padron):
    """{APELLIDO_sin_tildes: info} para fallback cuando el TSV trae solo apellido."""
    indice = {}
    for nombre_norm, info in padron.items():
        apellido = nombre_norm.split(",")[0].strip()
        clave = sin_tildes(apellido).upper()
        indice.setdefault(clave, info)
    return indice


def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    padron = construir_padron()
    with open(SALIDA, "w", encoding="utf-8") as f:
        json.dump(padron, f, ensure_ascii=False, indent=2)
    vigentes = sum(1 for v in padron.values() if v.get("vigente"))
    print(f"senadores.json -> {len(padron)} senadores ({vigentes} vigentes, "
          f"{len(padron) - vigentes} mandato cumplido)")


if __name__ == "__main__":
    main()
